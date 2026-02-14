# =============================================
# EXCEL SERVISI
# Tum moduller icin standart Excel export/import
# =============================================
#
# 📚 DERS: Bu servis her modülde ayni sekilde calisir.
# Yeni bir modul eklediginde (isyeri, calisan, ziyaret vs.)
# sadece alan_haritasi tanimlarsin, gerisi otomatik.
#
# Kullanim:
# excel_export(query, alan_haritasi) -> Excel dosyasi (bytes)
# excel_import(dosya, alan_haritasi, model) -> [{veri}, {veri}, ...]

from io import BytesIO
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


# 📚 DERS: Alan haritasi (field mapping)
# Her modul icin hangi alanlar Excel'de gorunecek, Turkce basliklari ne olacak
# ve hangi alanlar zorunlu - bunlari tanimlariz.
#
# Ornek:
# FIRMA_ALANLARI = [
#     {"alan": "ad",    "baslik": "Firma Adi",  "zorunlu": True,  "genislik": 30},
#     {"alan": "email", "baslik": "Email",      "zorunlu": True,  "genislik": 25},
# ]

FIRMA_ALANLARI = [
    {"alan": "ad",            "baslik": "Firma Adi",     "zorunlu": True,  "genislik": 30},
    {"alan": "kisa_ad",       "baslik": "Kisa Ad",       "zorunlu": False, "genislik": 15},
    {"alan": "il",            "baslik": "Il",            "zorunlu": True,  "genislik": 15},
    {"alan": "ilce",          "baslik": "Ilce",          "zorunlu": True,  "genislik": 15},
    {"alan": "email",         "baslik": "Email",         "zorunlu": True,  "genislik": 25},
    {"alan": "telefon",       "baslik": "Telefon",       "zorunlu": True,  "genislik": 18},
    {"alan": "adres",         "baslik": "Adres",         "zorunlu": False, "genislik": 40},
    {"alan": "vergi_dairesi", "baslik": "Vergi Dairesi", "zorunlu": False, "genislik": 20},
    {"alan": "vergi_no",      "baslik": "Vergi No",      "zorunlu": False, "genislik": 15},
]

ISYERI_ALANLARI = [
    {"alan": "ad",                "baslik": "Isyeri Adi",         "zorunlu": True,  "genislik": 30},
    {"alan": "firma_adi",         "baslik": "Firma Adi",          "zorunlu": False, "genislik": 25},
    {"alan": "sgk_sicil_no",      "baslik": "SGK Sicil No",       "zorunlu": True,  "genislik": 18},
    {"alan": "nace_kodu",         "baslik": "NACE Kodu",          "zorunlu": True,  "genislik": 12},
    {"alan": "nace_aciklama",     "baslik": "NACE Aciklama",      "zorunlu": False, "genislik": 30},
    {"alan": "tehlike_sinifi",    "baslik": "Tehlike Sinifi",      "zorunlu": True,  "genislik": 18},
    {"alan": "ana_faaliyet",      "baslik": "Ana Faaliyet",       "zorunlu": False, "genislik": 25},
    {"alan": "isveren_ad",        "baslik": "Isveren Adi",        "zorunlu": True,  "genislik": 18},
    {"alan": "isveren_soyad",     "baslik": "Isveren Soyadi",     "zorunlu": True,  "genislik": 18},
    {"alan": "isveren_vekili_ad", "baslik": "Vekil Adi",          "zorunlu": False, "genislik": 15},
    {"alan": "isveren_vekili_soyad", "baslik": "Vekil Soyadi",    "zorunlu": False, "genislik": 15},
    {"alan": "lokasyon",          "baslik": "Lokasyon",           "zorunlu": False, "genislik": 30},
]

CALISAN_ALANLARI = [
    {"alan": "ad",               "baslik": "Ad",              "zorunlu": True,  "genislik": 18},
    {"alan": "soyad",            "baslik": "Soyad",           "zorunlu": True,  "genislik": 18},
    {"alan": "tc_no",            "baslik": "TC No",           "zorunlu": False, "genislik": 15},
    {"alan": "telefon",          "baslik": "Telefon",         "zorunlu": False, "genislik": 18},
    {"alan": "email",            "baslik": "Email",           "zorunlu": False, "genislik": 25},
    {"alan": "dogum_tarihi",     "baslik": "Dogum Tarihi",    "zorunlu": False, "genislik": 15},
    {"alan": "ise_giris_tarihi", "baslik": "Ise Giris",       "zorunlu": False, "genislik": 15},
    {"alan": "gorev",            "baslik": "Gorev",           "zorunlu": False, "genislik": 20},
    {"alan": "bolum",            "baslik": "Bolum",           "zorunlu": False, "genislik": 20},
    {"alan": "kan_grubu",        "baslik": "Kan Grubu",       "zorunlu": False, "genislik": 12},
    {"alan": "isyeri_adi",       "baslik": "Isyeri",          "zorunlu": False, "genislik": 25},
]

PERSONEL_ALANLARI = [
    {"alan": "ad",                    "baslik": "Ad",                  "zorunlu": True,  "genislik": 18},
    {"alan": "soyad",                 "baslik": "Soyad",               "zorunlu": True,  "genislik": 18},
    {"alan": "unvan",                 "baslik": "Unvan",               "zorunlu": True,  "genislik": 20},
    {"alan": "tc_no",                 "baslik": "TC No",               "zorunlu": False, "genislik": 15},
    {"alan": "telefon",               "baslik": "Telefon",             "zorunlu": False, "genislik": 18},
    {"alan": "email",                 "baslik": "Email",               "zorunlu": False, "genislik": 25},
    {"alan": "uzmanlik_belgesi_no",   "baslik": "Belge No",            "zorunlu": False, "genislik": 18},
    {"alan": "diploma_no",            "baslik": "Diploma No",          "zorunlu": False, "genislik": 18},
    {"alan": "uzmanlik_sinifi",       "baslik": "Uzmanlik Sinifi",     "zorunlu": False, "genislik": 18},
    {"alan": "brans",                 "baslik": "Brans",               "zorunlu": False, "genislik": 25},
    {"alan": "ise_baslama_tarihi",    "baslik": "Ise Baslama",         "zorunlu": False, "genislik": 15},
]

# Ileride eklenecek moduller icin hazir:
# ZIYARET_ALANLARI = [...]


def excel_export(kayitlar: list, alan_haritasi: List[Dict], sayfa_adi: str = "Veriler") -> bytes:
    """
    📚 DERS: Veritabanindaki kayitlari Excel dosyasina cevirir.

    Parametreler:
    - kayitlar: Veritabanindan gelen ORM nesneleri listesi
    - alan_haritasi: Hangi alanlar Excel'e yazilacak
    - sayfa_adi: Excel sayfasinin adi

    Doner: Excel dosyasi (bytes olarak)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sayfa_adi

    # ---- STIL TANIMLARI ----
    baslik_font = Font(bold=True, color="FFFFFF", size=11)
    baslik_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    baslik_alignment = Alignment(horizontal="center", vertical="center")
    ince_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ---- BASLIK SATIRI ----
    for col, alan in enumerate(alan_haritasi, 1):
        hucre = ws.cell(row=1, column=col, value=alan["baslik"])
        hucre.font = baslik_font
        hucre.fill = baslik_fill
        hucre.alignment = baslik_alignment
        hucre.border = ince_border
        # Sutun genisligi
        ws.column_dimensions[hucre.column_letter].width = alan.get("genislik", 15)

    # ---- VERI SATIRLARI ----
    for row, kayit in enumerate(kayitlar, 2):
        for col, alan in enumerate(alan_haritasi, 1):
            # ORM nesnesinden veya dict'ten degeri al
            if isinstance(kayit, dict):
                deger = kayit.get(alan["alan"], "")
            else:
                deger = getattr(kayit, alan["alan"], "")

            hucre = ws.cell(row=row, column=col, value=deger)
            hucre.border = ince_border

    # ---- ALT BILGI ----
    son_satir = len(kayitlar) + 3
    ws.cell(row=son_satir, column=1, value=f"Toplam: {len(kayitlar)} kayit")
    ws.cell(row=son_satir + 1, column=1, value=f"Olusturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    # Bytes olarak dondur
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def excel_sablon_olustur(alan_haritasi: List[Dict], sayfa_adi: str = "Sablon") -> bytes:
    """
    📚 DERS: Bos Excel sablonu olusturur (iceri aktarim icin).

    Kullanici bu sablonu indirir, doldurur ve geri yukler.
    Zorunlu alanlar kirmizi baslikla isaretlenir.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sayfa_adi

    # Stiller
    zorunlu_font = Font(bold=True, color="FFFFFF", size=11)
    zorunlu_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
    normal_font = Font(bold=True, color="FFFFFF", size=11)
    normal_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    baslik_alignment = Alignment(horizontal="center", vertical="center")

    # Baslik satiri
    for col, alan in enumerate(alan_haritasi, 1):
        baslik = alan["baslik"]
        if alan.get("zorunlu"):
            baslik += " *"

        hucre = ws.cell(row=1, column=col, value=baslik)
        hucre.font = zorunlu_font if alan.get("zorunlu") else normal_font
        hucre.fill = zorunlu_fill if alan.get("zorunlu") else normal_fill
        hucre.alignment = baslik_alignment
        ws.column_dimensions[hucre.column_letter].width = alan.get("genislik", 15)

    # Aciklama satiri (2. satir)
    aciklama_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    for col, alan in enumerate(alan_haritasi, 1):
        aciklama = "Zorunlu" if alan.get("zorunlu") else "Opsiyonel"
        hucre = ws.cell(row=2, column=col, value=aciklama)
        hucre.fill = aciklama_fill
        hucre.font = Font(italic=True, size=9, color="666666")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def excel_import(dosya_icerik: bytes, alan_haritasi: List[Dict]) -> Dict[str, Any]:
    """
    📚 DERS: Excel dosyasindan veri okur ve dogrular.

    Doner:
    {
        "basarili": [{veri1}, {veri2}, ...],
        "hatali": [{"satir": 3, "hata": "Email zorunlu"}, ...],
        "toplam": 10,
        "basarili_sayisi": 8,
        "hatali_sayisi": 2
    }
    """
    try:
        wb = load_workbook(BytesIO(dosya_icerik), read_only=True)
    except Exception:
        return {
            "basarili": [],
            "hatali": [{"satir": 0, "hata": "Gecersiz Excel dosyasi. Lutfen .xlsx formatinda yukleyin."}],
            "toplam": 0,
            "basarili_sayisi": 0,
            "hatali_sayisi": 1,
        }

    ws = wb.active

    # ---- BASLIK SATIRINI OKU ----
    # Excel'deki baslik sirasi ile alan_haritasi esle
    basliklar = []
    for col in range(1, ws.max_column + 1):
        deger = ws.cell(row=1, column=col).value
        if deger:
            # "Firma Adi *" -> "Firma Adi" (yildizi temizle)
            basliklar.append(str(deger).replace(" *", "").strip())
        else:
            basliklar.append("")

    # Baslik -> alan adi eslestirmesi
    baslik_alan_map = {}
    for alan in alan_haritasi:
        for i, baslik in enumerate(basliklar):
            if baslik == alan["baslik"]:
                baslik_alan_map[i] = alan
                break

    if not baslik_alan_map:
        return {
            "basarili": [],
            "hatali": [{"satir": 0, "hata": "Excel basliklari taninamadi. Lutfen sablonu kullanin."}],
            "toplam": 0,
            "basarili_sayisi": 0,
            "hatali_sayisi": 1,
        }

    # ---- VERI SATIRLARINI OKU ----
    basarili = []
    hatali = []
    baslangic_satir = 2  # 1. satir baslik

    # 2. satir "Zorunlu/Opsiyonel" aciklama satiri olabilir - atla
    ikinci_satir = ws.cell(row=2, column=1).value
    if ikinci_satir and str(ikinci_satir).strip() in ("Zorunlu", "Opsiyonel"):
        baslangic_satir = 3

    for row in range(baslangic_satir, ws.max_row + 1):
        # Bos satiri atla
        tum_bos = True
        for col in range(ws.max_column):
            if ws.cell(row=row, column=col + 1).value is not None:
                tum_bos = False
                break
        if tum_bos:
            continue

        kayit = {}
        satir_hatalari = []

        for col_index, alan_bilgi in baslik_alan_map.items():
            deger = ws.cell(row=row, column=col_index + 1).value
            alan_adi = alan_bilgi["alan"]

            # None -> bos string
            if deger is not None:
                deger = str(deger).strip()
            else:
                deger = ""

            # Zorunlu alan kontrolu
            if alan_bilgi.get("zorunlu") and not deger:
                satir_hatalari.append(f"{alan_bilgi['baslik']} zorunlu")

            kayit[alan_adi] = deger if deger else None

        if satir_hatalari:
            hatali.append({
                "satir": row,
                "hata": ", ".join(satir_hatalari),
                "veri": kayit,
            })
        else:
            basarili.append(kayit)

    wb.close()

    return {
        "basarili": basarili,
        "hatali": hatali,
        "toplam": len(basarili) + len(hatali),
        "basarili_sayisi": len(basarili),
        "hatali_sayisi": len(hatali),
    }
